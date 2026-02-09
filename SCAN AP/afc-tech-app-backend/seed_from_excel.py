# seed_from_excel.py
import os
import re
import sys
from datetime import datetime, date

import pandas as pd
import openpyxl

from app import app
from db import db
from models import Hospital, AHU, Filter, Building


EXCEL_PATH = "./excel_data_raw/filter-datasheet.xlsm"


# -----------------------------
# Helpers
# -----------------------------
def clean_str(x):
    """Trim strings safely; treat NaN/None/empty as None."""
    if x is None:
        return None
    try:
        if pd.isna(x):
            return None
    except Exception:
        pass
    s = str(x).strip()
    if not s:
        return None
    # normalize common NaN-ish strings
    if s.strip().lower() in ("nan", "none", "null", "n/a", "na", "-", "empty"):
        return None
    return s


def is_placeholder(s):
    if not s:
        return True
    normalized = str(s).strip().lower()
    return normalized in ["empty", "nan", "n/a", "na", "-", "none", "null"]


def to_date(val):
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except Exception:
        pass

    if isinstance(val, pd.Timestamp):
        return val.date()
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return pd.to_datetime(val).date()
    except Exception:
        return None


def parse_quantity(val, default=1):
    if val is None:
        return default
    try:
        if pd.isna(val):
            return default
    except Exception:
        pass

    if isinstance(val, (int, float)):
        try:
            return int(val)
        except Exception:
            return default

    s = str(val).strip()
    if not s:
        return default

    m = re.search(r"\d+", s)
    if m:
        try:
            return int(m.group(0))
        except Exception:
            return default

    return default


def parse_frequency_to_days(raw):
    if raw is None:
        return None

    s = str(raw).strip()
    if not s:
        return None

    if s.lower() == "removed":
        return None

    m = re.search(r"(\d+)\s*day", s, re.IGNORECASE)
    if m:
        return int(m.group(1))

    m = re.search(r"(\d+)\s*month", s, re.IGNORECASE)
    if m:
        return int(m.group(1)) * 30

    m = re.search(r"(\d+)\s*year", s, re.IGNORECASE)
    if m:
        return int(m.group(1)) * 365

    return None


def get_sheet_title_cell(path, sheet_name="MAIN BUILDING", cell="B2"):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    return clean_str(ws[cell].value)


def has_attr(obj, attr: str) -> bool:
    return hasattr(obj, attr)


def normalize_ahu_key(display_name: str, building: str = None) -> str:
    """
    Logical grouping key so multiple filter rows map to the same AHU.
    Include building so 'AHU-1' in different buildings won't collide.
    """
    dn = clean_str(display_name) or ""
    b = clean_str(building) or ""
    raw = f"{b}::{dn}".strip().lower()
    raw = re.sub(r"\s+", " ", raw).strip()
    return raw or "unnamed"


def make_sequential_ahu_id(seq: int) -> str:
    return f"AHU-{seq:03d}"


# -----------------------------
# Upserts
# -----------------------------
def upsert_hospital(name: str):
    h = Hospital.query.filter_by(name=name).first()
    if h:
        if hasattr(h, "active") and h.active is None:
            h.active = True
        return h

    h = Hospital(name=name, active=True)
    db.session.add(h)
    db.session.flush()
    return h


def upsert_building(hospital_id: int, name: str, floor_area: str = None):
    name = clean_str(name)
    if not name:
        return None

    b = Building.query.filter_by(hospital_id=hospital_id, name=name).first()
    if b:
        if floor_area and has_attr(b, "floor_area") and not b.floor_area:
            b.floor_area = floor_area
        return b

    b = Building(hospital_id=hospital_id, name=name, floor_area=floor_area, active=True)
    db.session.add(b)
    db.session.flush()
    return b


def upsert_ahu(
    ahu_id: str,
    hospital_id: int,
    display_name: str = None,
    location=None,
    notes=None,
    excel_order=None,
    building_id=None
):
    # Accept either an integer AHU id or a logical label. When seeding
    # we generally create AHU records directly and then map logical keys
    # to the assigned integer id. This helper will try to fetch by id
    # when possible; otherwise it will return None (seeding logic will
    # create new AHU records explicitly).
    if ahu_id is None:
        return None
    try:
        aid = int(ahu_id)
    except Exception:
        return None

    a = db.session.get(AHU, aid)
    if a:
        a.hospital_id = hospital_id
        if building_id is not None and has_attr(a, "building_id"):
            a.building_id = building_id
        if display_name:
            a.name = display_name
        a.location = clean_str(location) or a.location
        a.notes = clean_str(notes) or a.notes

        if excel_order is not None and has_attr(a, "excel_order"):
            a.excel_order = int(excel_order)
        return a

    return None


def find_existing_filter(ahu_id, phase, part_number, size):
    return (
        Filter.query.filter_by(
            ahu_id=ahu_id,
            phase=phase,
            part_number=part_number,
            size=size,
        ).first()
    )


def upsert_filter(
    ahu_id,
    phase,
    part_number,
    size,
    quantity,
    frequency_days,
    last_service_date,
    is_active=True,
    excel_order=None,
):
    # ahu_id may be an integer (preferred) or a label string; support both
    if isinstance(ahu_id, str):
        ahu_id_clean = clean_str(ahu_id)
        if not ahu_id_clean:
            return None
        try:
            ahu_id_val = int(ahu_id_clean)
        except Exception:
            # if it's not numeric, abort (we expect numeric AHU IDs now)
            return None
    else:
        ahu_id_val = ahu_id

    phase = clean_str(phase)
    part_number = clean_str(part_number) or ""
    size = clean_str(size)

    if not ahu_id_val or not size:
        return None
    existing = find_existing_filter(ahu_id_val, phase, part_number, size)
    if existing:
        try:
            existing.quantity = parse_quantity(quantity, default=getattr(existing, "quantity", 1) or 1)
        except Exception:
            pass

        if frequency_days is not None:
            try:
                existing.frequency_days = int(frequency_days)
            except Exception:
                pass

        if last_service_date:
            existing.last_service_date = last_service_date

        existing.is_active = bool(is_active)

        if excel_order is not None and has_attr(existing, "excel_order"):
            existing.excel_order = int(excel_order)

        return existing

    if frequency_days is None:
        frequency_days = 90

    kwargs = dict(
        ahu_id=ahu_id_val,
        phase=phase,
        part_number=part_number,
        size=size,
        quantity=parse_quantity(quantity),
        frequency_days=int(frequency_days),
        last_service_date=last_service_date,
        is_active=bool(is_active),
    )

    if excel_order is not None and has_attr(Filter, "excel_order"):
        kwargs["excel_order"] = int(excel_order)

    f = Filter(**kwargs)
    db.session.add(f)
    return f


# -----------------------------
# Main seed
# -----------------------------
def seed_from_excel(path, selected_sheet=None):
    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel file not found: {path}")

    xls = pd.ExcelFile(path)

    if selected_sheet:
        if selected_sheet not in xls.sheet_names:
            raise ValueError(f"Sheet '{selected_sheet}' not found. Available: {xls.sheet_names}")
        data_sheets = [selected_sheet]
        preferred_sheet = selected_sheet
    else:
        data_sheets = [s for s in xls.sheet_names if s.strip().lower() != "filter"]
        if not data_sheets:
            raise RuntimeError("No data sheets found.")
        preferred_sheet = "MAIN BUILDING" if "MAIN BUILDING" in data_sheets else data_sheets[0]

    hospital_name = get_sheet_title_cell(path, sheet_name=preferred_sheet, cell="B2") or preferred_sheet.upper().replace("_", " ")
    hospital = upsert_hospital(hospital_name)

    stats = {
        "hospital": hospital_name,
        "sheets_processed": 0,
        "rows_seen": 0,
        "ahus": set(),
        "filters_upserted": 0,
        "filters_skipped": 0,
    }

    # logical key -> sequential id
    ahu_key_to_id = {}
    next_seq = 1

    for sheet in data_sheets:
        df = pd.read_excel(path, sheet_name=sheet, header=4)
        df.columns = [str(c).strip() for c in df.columns]

        def col(exact_name):
            target = re.sub(r"\s+", " ", exact_name).strip().lower()
            for c in df.columns:
                c_norm = re.sub(r"\s+", " ", str(c)).strip().lower()
                if c_norm == target:
                    return c
            return None

        col_ahu = col("AHU NO.")
        col_loc = col("LOCATION")
        col_stage = col("STAGE")
        col_size = col("FILTER SIZE")
        col_freq = col("FREQUENCY")
        col_qty = col("QUANTITY") or col("QTY")

        col_building = col("BUILDING")
        col_floor_area = col("FLOOR/AREA")
        col_part_num = col("PART NUMBER")
        col_filter_type = col("FILTER TYPE")
        col_repl = col("DATE OF REPLACEMENT")

        required = [col_ahu, col_loc, col_stage, col_size, col_qty, col_freq]
        if any(x is None for x in required):
            print(f"Skipping sheet '{sheet}' (missing required columns). Found columns: {list(df.columns)}")
            continue

        stats["sheets_processed"] += 1
        stats["rows_seen"] += len(df)

        filter_order_map = {}

        last_display_name = None
        last_row_had_ahu = False

        for _, r in df.iterrows():
            raw_ahu_no = clean_str(r.get(col_ahu))
            location = clean_str(r.get(col_loc)) if col_loc else None
            building = clean_str(r.get(col_building)) if col_building else None
            floor_area = clean_str(r.get(col_floor_area)) if col_floor_area else None

            # Determine whether this row contains any filter-related data
            has_filter_data = False
            for c in (col_stage, col_size, col_qty, col_part_num, col_filter_type, col_freq, col_repl):
                if c:
                    val = clean_str(r.get(c))
                    if val and not is_placeholder(val):
                        has_filter_data = True
                        break

            # Decide display name
            if not is_placeholder(raw_ahu_no):
                display_name = raw_ahu_no
                last_display_name = display_name
                last_row_had_ahu = True
            else:
                if has_filter_data:
                    if last_row_had_ahu and last_display_name:
                        display_name = last_display_name
                    else:
                        display_name = f"Unnamed — {location}" if location else f"Unnamed — {next_seq}"
                        last_display_name = display_name
                        last_row_had_ahu = True
                else:
                    display_name = None
                    last_display_name = None
                    last_row_had_ahu = False

            if display_name is None:
                continue

            ahu_key = normalize_ahu_key(display_name, building=building)

            # ensure building exists before creating AHU so we can set building_id
            building_obj = None
            building_id = None
            if building:
                building_obj = upsert_building(hospital.id, building, floor_area=floor_area)
                building_id = building_obj.id if building_obj else None

            if ahu_key not in ahu_key_to_id:
                # create a new AHU row; DB will assign a numeric id
                display_label = make_sequential_ahu_id(next_seq)
                a = AHU(
                    hospital_id=hospital.id,
                    building_id=building_id,
                    name=display_name or display_label,
                    location=location,
                    notes=None,
                    excel_order=next_seq,
                )
                db.session.add(a)
                db.session.flush()  # assign id
                ahu_id = a.id
                ahu_key_to_id[ahu_key] = ahu_id
                excel_order = next_seq
                next_seq += 1
            else:
                ahu_id = ahu_key_to_id[ahu_key]
                # try to fetch existing excel_order if present
                existing_ahu = db.session.get(AHU, ahu_id)
                excel_order = getattr(existing_ahu, "excel_order", None) if existing_ahu else None
            notes_parts = [
                f"Excel AHU: {raw_ahu_no or '(blank)'}",
                f"Excel Display: {display_name}",
            ]
            if building:
                notes_parts.append(f"Building: {building}")
            if floor_area:
                notes_parts.append(f"Floor/Area: {floor_area}")
            notes = " | ".join(notes_parts)

            # Ensure AHU object exists and update notes/excel_order if needed
            ahu_obj = db.session.get(AHU, ahu_id)
            if ahu_obj:
                ahu_obj.notes = notes or ahu_obj.notes
                if excel_order is not None and hasattr(ahu_obj, "excel_order"):
                    ahu_obj.excel_order = int(excel_order)
            stats["ahus"].add(ahu_id)

            # filters
            phase = clean_str(r.get(col_stage))
            size = clean_str(r.get(col_size))
            qty = r.get(col_qty)

            freq_raw = r.get(col_freq)
            freq_days = parse_frequency_to_days(freq_raw)

            part_number = clean_str(r.get(col_part_num)) if col_part_num else None
            if not part_number:
                part_number = clean_str(r.get(col_filter_type)) if col_filter_type else None

            last_service_date = to_date(r.get(col_repl)) if col_repl else None

            is_active = True
            if isinstance(freq_raw, str) and freq_raw.strip().lower() == "removed":
                is_active = False

            if not size:
                stats["filters_skipped"] += 1
                continue

            filter_order_map.setdefault(ahu_id, 1)
            filter_excel_order = filter_order_map[ahu_id]
            filter_order_map[ahu_id] += 1

            upsert_filter(
                ahu_id=ahu_id,
                phase=phase,
                part_number=part_number,
                size=size,
                quantity=qty,
                frequency_days=freq_days,
                last_service_date=last_service_date,
                is_active=is_active,
                excel_order=filter_excel_order,
            )
            stats["filters_upserted"] += 1

    db.session.commit()

    print("\n✅ Seed complete")
    print(f"Hospital: {stats['hospital']} (ID {hospital.id})")
    print(f"Sheets processed: {stats['sheets_processed']}")
    print(f"Rows seen: {stats['rows_seen']}")
    print(f"AHUs upserted: {len(stats['ahus'])}")
    print(f"Filters upserted: {stats['filters_upserted']}")
    print(f"Filters skipped (missing size): {stats['filters_skipped']}")

    ordered = sorted(list(stats["ahus"]))
    print("\nExample AHU IDs (first 15):")
    for i, x in enumerate(ordered[:15], start=1):
        print(f"  {i}. AHU-{int(x):03d} (db id: {x})")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python seed_from_excel.py <sheet_name>")
        print("Available sheets:")
        xls = pd.ExcelFile(EXCEL_PATH)
        for sheet in xls.sheet_names:
            print(f"  - {sheet}")
        sys.exit(1)

    selected_sheet = sys.argv[1]
    with app.app_context():
        seed_from_excel(EXCEL_PATH, selected_sheet)
